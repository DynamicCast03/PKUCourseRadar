import json
import re
from openpyxl import load_workbook
from collections import defaultdict

def parse_excel(file_path, day_index):
    wb = load_workbook(file_path, read_only=False, data_only=True) 
    courses = []
    for sheet in wb:
        building = sheet.title  
        merged_ranges = list(sheet.merged_cells.ranges)
        rows = list(sheet.iter_rows(values_only=True))
        if len(rows) < 3: 
            continue
        periods_row = rows[0]  
        times_row = rows[1]   
        period_mapping = parse_periods(periods_row)
        if not period_mapping:
            print(f"警告: 在工作表 '{building}' 中未找到有效的节次信息")
            continue
        for row_idx, row in enumerate(rows[2:], start=3):
            if not any(row): 
                continue
            room = str(row[0]).strip() if row[0] else ""
            if not room or room in ['None', '']:
                continue
            parsed_courses = parse_row_courses(row, period_mapping, building, room, day_index, merged_ranges, row_idx)
            courses.extend(parsed_courses)
    return courses

def parse_periods(periods_row):
    period_mapping = {} 
    chinese_nums = {
        '一': 1, '二': 2, '三': 3, '四': 4, '五': 5, '六': 6, '七': 7, '八': 8,
        '九': 9, '十': 10, '十一': 11, '十二': 12, '十三': 13, '十四': 14
    }
    for col_idx, period_cell in enumerate(periods_row[1:], start=1):  
        if not period_cell:
            continue 
        period_str = str(period_cell).strip()
        period_num = chinese_nums.get(period_str, 0)
        if period_num > 0:
            period_mapping[col_idx] = period_num
    return period_mapping

def parse_row_courses(row, period_mapping, building, room, day_index, merged_ranges, row_idx):
    courses = []
    processed_cols = set() 
    for col_idx, cell_content in enumerate(row[1:], start=1):
        if col_idx in processed_cols:
            continue
        if not cell_content or str(cell_content).strip() in ['', 'None']:
            continue
        cell_str = str(cell_content).strip()
        if col_idx not in period_mapping:
            continue
        course_periods = []
        merged_cols = get_merged_columns(merged_ranges, row_idx, col_idx)
        if merged_cols:
            for merged_col in merged_cols:
                if merged_col in period_mapping:
                    course_periods.append(period_mapping[merged_col])
                    processed_cols.add(merged_col)
        else:
            current_col = col_idx
            while current_col < len(row) and current_col in period_mapping:
                if current_col < len(row) and row[current_col] and str(row[current_col]).strip() == cell_str:
                    course_periods.append(period_mapping[current_col])
                    processed_cols.add(current_col)
                    current_col += 1
                else:
                    break
        if not course_periods:
            course_periods.append(period_mapping[col_idx])
            processed_cols.add(col_idx)
        course_periods = ensure_minimum_periods(course_periods, period_mapping, col_idx)
        parsed_courses = parse_course_cell(cell_str, building, room, day_index, course_periods)
        courses.extend(parsed_courses)
    return courses

def get_merged_columns(merged_ranges, row_idx, col_idx):
    for merged_range in merged_ranges:
        if (merged_range.min_row <= row_idx <= merged_range.max_row and 
            merged_range.min_col <= col_idx + 1 <= merged_range.max_col): 
            return list(range(merged_range.min_col - 1, merged_range.max_col))
    return []

def ensure_minimum_periods(periods, period_mapping, start_col):
    if len(periods) >= 2:
        return sorted(periods)
    if len(periods) == 1:
        current_period = periods[0]
        next_col = start_col + 1
        if next_col in period_mapping:
            next_period = period_mapping[next_col]
            if next_period == current_period + 1: 
                periods.append(next_period)
    return sorted(periods)

def parse_course_cell(cell_content, building, room, day_index, periods):
    courses = []
    course_lines = re.split(r'[\n\r]+', cell_content)
    for line in course_lines:
        line = line.strip()
        if not line:
            continue
        course_info = parse_course_line(line)
        if course_info:
            course = create_course_object(
                course_info['name'],
                course_info['teacher'],
                periods,  
                building,
                room,
                day_index
            )
            courses.append(course)
    return courses

def parse_course_line(line):
    match = re.match(r'([^（）]+)（([^）]+)）', line)
    if match:
        name = match.group(1).strip()
        teacher = match.group(2).strip()
        return {
            'name': name,
            'teacher': teacher
        }
    match = re.match(r'([^[\]]+)\[([^\]]+)\]', line)
    if match:
        name = match.group(1).strip()
        teacher = match.group(2).strip()
        return {
            'name': name,
            'teacher': teacher
        }
    return {
        'name': line.strip(),
        'teacher': "未知教师"
    }

def create_course_object(name, teacher, periods, building, room, day_index):
    course_time = [[] for _ in range(7)]
    if 0 <= day_index < 7:
        course_time[day_index] = sorted(periods)  
    tags = generate_tags(name)
    return {
        "name": name,
        "courseTime": course_time,
        "tags": tags,
        "building": building,
        "room": room,
        "teacher": teacher
    }

def generate_tags(course_name):
    tags = []
    subject_keywords = {
        "数学": ["数学", "微积分", "线性代数", "概率", "统计", "几何", "高等数学", "数理", "数分", "高数"],
        "计算机": ["计算机", "编程", "程序设计", "算法", "数据结构", "软件", "网络", "数据库", "计算", "程序"],
        "人工智能": ["人工智能", "机器学习", "神经网络", "深度学习", "AI", "智能"],
        "物理": ["物理", "力学", "电磁", "量子", "热力学", "普通物理", "大学物理"],
        "化学": ["化学", "有机", "无机", "分析", "物化", "化工", "材料化学"],
        "生物": ["生物", "遗传", "细胞", "分子", "生态", "生命科学", "生物学"],
        "经济管理": ["经济", "金融", "会计", "管理", "市场", "经济学", "商务", "企业管理"],
        "语言文学": ["英语", "汉语", "日语", "语言", "文学", "口语", "听力", "精读", "写作", "翻译", "文化"],
        "工程技术": ["工程", "机械", "电子", "建筑", "材料", "技术", "工程学", "制造"],
        "政治法律": ["政治", "法学", "法律", "法理", "宪法", "民法", "思想", "马克思", "政府", "公共政策"],
        "哲学": ["哲学", "逻辑", "伦理", "美学", "思维", "哲学史"],
        "历史": ["历史", "史学", "古代", "近代", "现代", "中国通史", "世界史", "文明史"],
        "地理环境": ["地理", "地质", "环境", "气候", "地球", "环境科学", "生态环境"],
        "体育健康": ["体育", "运动", "健身", "球类", "健康", "体能", "武术"],
        "艺术设计": ["艺术", "美术", "音乐", "设计", "绘画", "创意", "视觉", "艺术设计"],
        "医学": ["医学", "临床", "病理", "解剖", "药理", "医疗", "健康科学"],
        "心理学": ["心理", "心理学", "行为", "认知", "心理健康"],
        "社会学": ["社会", "社会学", "人类学", "民族", "社会科学", "文化研究"]
    }
    
    for subject, keywords in subject_keywords.items():
        if any(keyword in course_name for keyword in keywords):
            tags.append(subject)
    if any(keyword in course_name for keyword in ["实验", "实习", "实训", "实践"]):
        tags.append("实践课程")
    if any(keyword in course_name for keyword in ["论坛", "讲座", "报告"]):
        tags.append("讲座论坛")
    if any(keyword in course_name for keyword in ["讨论", "研讨", "交流"]):
        tags.append("研讨交流")
    if any(keyword in course_name for keyword in ["设计", "创作", "制作"]):
        tags.append("创作设计")
    if any(keyword in course_name for keyword in ["导论", "概论", "入门", "基础"]):
        tags.append("入门导论")
    if any(keyword in course_name for keyword in ["高级", "高等", "进阶", "深度", "专题"]):
        tags.append("高级进阶")
    if any(keyword in course_name for keyword in ["理论", "原理", "学说"]):
        tags.append("理论课程")
    if any(keyword in course_name for keyword in ["应用", "实用", "技能"]):
        tags.append("应用技能")
    if any(keyword in course_name for keyword in ["分析", "研究", "方法"]):
        tags.append("分析研究")
    if any(keyword in course_name for keyword in ["历史", "发展", "演变"]):
        tags.append("历史发展")
    if any(keyword in course_name for keyword in ["国际", "世界", "全球"]):
        tags.append("国际视野")
    if any(keyword in course_name for keyword in ["中国", "中华", "本土"]):
        tags.append("中国特色")
    if any(keyword in course_name for keyword in ["现代", "当代", "新"]):
        tags.append("现代前沿")
    if any(keyword in course_name for keyword in ["传统", "古典", "经典"]):
        tags.append("传统经典")
    if "班" in course_name:
        tags.append("班级教学")
    if any(keyword in course_name for keyword in ["个人", "独立", "自主"]):
        tags.append("自主学习")
    if any(keyword in course_name for keyword in ["团队", "合作", "协作"]):
        tags.append("合作学习")
    if any(keyword in course_name for keyword in ["在线", "网络", "远程"]):
        tags.append("在线课程")
    if "A" in course_name or "B" in course_name or "C" in course_name:
        tags.append("分级课程")
    if "I" in course_name or "II" in course_name or "III" in course_name or "（一）" in course_name or "（二）" in course_name or "（三）" in course_name:
        tags.append("系列课程")
    if any(keyword in course_name for keyword in ["选修", "任选", "选择"]):
        tags.append("选修课程")
    if any(keyword in course_name for keyword in ["必修", "核心", "主干"]):
        tags.append("核心课程")
    if not tags:
        tags.append("通识课程")
    return list(set(tags)) 

def merge_same_courses(courses):
    course_dict = {}
    for course in courses:
        key = (course['name'], course['teacher'], course['building'], course['room'])
        if key in course_dict:
            for day_idx in range(7):
                existing_periods = set(course_dict[key]['courseTime'][day_idx])
                new_periods = set(course['courseTime'][day_idx])
                merged_periods = sorted(list(existing_periods | new_periods))
                course_dict[key]['courseTime'][day_idx] = merged_periods
        else:
            course_dict[key] = course.copy()
    return list(course_dict.values())

def get_day_index_from_filename(filename):
    try:
        match = re.search(r'(\d+)', filename)
        if match:
            day_num = int(match.group(1))
            return day_num - 1
    except:
        pass
    return 0 

def main():
    excel_files = ["1.xlsx", "2.xlsx", "3.xlsx", "4.xlsx", "5.xlsx"]
    all_courses = []
    for file_name in excel_files:
        try:
            print(f"正在处理文件: {file_name}")
            day_index = get_day_index_from_filename(file_name)
            day_names = ["周一", "周二", "周三", "周四", "周五", "周六", "周日"]
            print(f"  -> 处理 {day_names[day_index]} 的课程")
            courses = parse_excel(file_name, day_index)
            all_courses.extend(courses)
            print(f"  -> 从 {file_name} 提取了 {len(courses)} 个课程时间段")
        except FileNotFoundError:
            print(f"文件 {file_name} 不存在，跳过")
        except Exception as e:
            print(f"处理文件 {file_name} 时发生错误: {str(e)}")
            import traceback
            traceback.print_exc()
    merged_courses = merge_same_courses(all_courses)
    output_file = "courses.json"
    with open(output_file, "w", encoding="utf-8") as f:
        json.dump(merged_courses, f, indent=2, ensure_ascii=False)
    print(f"\n处理完成！")
    print(f"总共提取了 {len(all_courses)} 个课程时间段")
    print(f"合并后共有 {len(merged_courses)} 门不同的课程")
    print(f"结果已保存到 {output_file}")
    if merged_courses:
        print(f"\n前3门课程示例:")
        day_names = ["周一", "周二", "周三", "周四", "周五", "周六", "周日"]
        for i, course in enumerate(merged_courses[:3]):
            print(f"{i+1}. {course['name']}")
            print(f"   教师: {course['teacher']}")
            print(f"   地点: {course['building']} {course['room']}")
            print(f"   标签: {', '.join(course['tags'])}")
            schedule = []
            for day_idx, periods in enumerate(course['courseTime']):
                if periods:
                    schedule.append(f"{day_names[day_idx]}第{','.join(map(str, periods))}节")
            print(f"   时间: {'; '.join(schedule) if schedule else '无'}")
            print()
if __name__ == "__main__":
    main()